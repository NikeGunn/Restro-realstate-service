"""
ExcelParser
===========
Two-phase parser: parse() reads + validates rows without DB writes; the
caller (Celery task or view) decides whether to commit. Column detection
uses fuzzy matching so headers like "Item Name" and "product" both map.

Supports .xlsx (openpyxl), .xls (xlrd), .csv (built-in csv).
"""
import csv
import difflib
import os
from datetime import date, datetime
from decimal import Decimal, InvalidOperation


COLUMN_SYNONYMS = {
    'name': ['name', 'item_name', 'item name', 'product', 'product name',
             'description', 'item', 'material'],
    'quantity': ['quantity', 'qty', 'amount', 'count', 'units', 'vol', 'volume'],
    'unit_cost': ['unit_cost', 'unit cost', 'unit_price', 'unit price', 'cost',
                  'price', 'rate', 'price per unit', 'wholesale'],
    'supplier_name': ['supplier', 'vendor', 'vendor name', 'supplier name', 'from'],
    'sku': ['sku', 'code', 'item_code', 'item code', 'product_code',
            'ref', 'reference', 'barcode'],
    'movement_date': ['date', 'order_date', 'order date', 'invoice_date',
                      'invoice date', 'sale_date', 'sale date',
                      'transaction_date', 'transaction date'],
    'category': ['category', 'type', 'group', 'dept', 'department'],
    'notes': ['notes', 'remarks', 'comment', 'memo'],
    'unit': ['unit', 'uom', 'measure', 'unit of measure'],
}


def _to_decimal(val) -> Decimal:
    if val is None or val == '':
        return Decimal('0')
    try:
        return Decimal(str(val).replace(',', '').strip())
    except (InvalidOperation, ValueError):
        return Decimal('0')


def _parse_date(val):
    if val is None or val == '':
        return None
    if isinstance(val, (date, datetime)):
        return val if isinstance(val, date) and not isinstance(val, datetime) else val.date()
    s = str(val).strip()
    fmts = (
        '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y',
        '%Y/%m/%d', '%d.%m.%Y', '%m-%d-%Y',
    )
    for f in fmts:
        try:
            return datetime.strptime(s, f).date()
        except ValueError:
            continue
    # Last-resort lenient parse via dateutil if available
    try:
        from dateutil import parser as du
        return du.parse(s).date()
    except Exception:
        return None


class ExcelParser:
    IMPORT_TYPES = ('sales', 'purchase', 'supplier_catalog')
    ERROR_ABORT_THRESHOLD = Decimal('0.30')  # >30% errors → caller marks failed

    def __init__(self, file_path: str, import_type: str, organization, location=None):
        if import_type not in self.IMPORT_TYPES:
            raise ValueError(f'Unknown import_type: {import_type}')
        self.file_path = file_path
        self.import_type = import_type
        self.organization = organization
        self.location = location

    # ──────────────────────────────────────────────────────────────
    # Reading
    # ──────────────────────────────────────────────────────────────
    def _read_rows(self):
        ext = os.path.splitext(self.file_path)[1].lower()
        if ext == '.csv':
            return list(self._read_csv())
        if ext == '.xlsx':
            return list(self._read_xlsx())
        if ext == '.xls':
            return list(self._read_xls())
        raise ValueError(f'Unsupported file extension: {ext}')

    def _read_csv(self):
        with open(self.file_path, newline='', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            for row in reader:
                yield row

    def _read_xlsx(self):
        from openpyxl import load_workbook
        wb = load_workbook(self.file_path, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            yield list(row)
        wb.close()

    def _read_xls(self):
        try:
            import xlrd  # noqa
        except ImportError:
            raise ValueError('.xls support requires xlrd to be installed.')
        book = xlrd.open_workbook(self.file_path)
        sheet = book.sheet_by_index(0)
        for r in range(sheet.nrows):
            yield [sheet.cell_value(r, c) for c in range(sheet.ncols)]

    # ──────────────────────────────────────────────────────────────
    # Column detection
    # ──────────────────────────────────────────────────────────────
    def detect_columns(self, headers: list) -> dict:
        result = {}
        cleaned = [(h or '').strip().lower() for h in headers]
        for standard, synonyms in COLUMN_SYNONYMS.items():
            best_idx = None
            for i, h in enumerate(cleaned):
                if not h:
                    continue
                if h in synonyms:
                    best_idx = i
                    break
                if difflib.get_close_matches(h, synonyms, n=1, cutoff=0.75):
                    best_idx = i
                    break
            if best_idx is not None:
                result[standard] = best_idx
        return result

    # ──────────────────────────────────────────────────────────────
    # Row validation
    # ──────────────────────────────────────────────────────────────
    def validate_row(self, row_data: dict, row_num: int) -> list:
        errors = []
        if self.import_type == 'sales':
            qty = _to_decimal(row_data.get('quantity'))
            if qty <= 0:
                errors.append({'row': row_num, 'column': 'quantity',
                               'message': 'Positive quantity is required.'})
            if not row_data.get('movement_date'):
                errors.append({'row': row_num, 'column': 'movement_date',
                               'message': 'Date is required.'})
            if not (row_data.get('name') or row_data.get('sku')):
                errors.append({'row': row_num, 'column': 'name/sku',
                               'message': 'Item name or SKU is required.'})
        elif self.import_type == 'purchase':
            qty = _to_decimal(row_data.get('quantity'))
            if qty <= 0:
                errors.append({'row': row_num, 'column': 'quantity',
                               'message': 'Positive quantity is required.'})
            if _to_decimal(row_data.get('unit_cost')) <= 0:
                errors.append({'row': row_num, 'column': 'unit_cost',
                               'message': 'Positive unit cost is required.'})
            if not (row_data.get('name') or row_data.get('sku')):
                errors.append({'row': row_num, 'column': 'name/sku',
                               'message': 'Item name or SKU is required.'})
        elif self.import_type == 'supplier_catalog':
            if not row_data.get('name'):
                errors.append({'row': row_num, 'column': 'name',
                               'message': 'Item name is required.'})
        return errors

    # ──────────────────────────────────────────────────────────────
    # Item resolution
    # ──────────────────────────────────────────────────────────────
    def resolve_item(self, row_data: dict):
        from apps.inventory.models import InventoryItem
        sku = (row_data.get('sku') or '').strip()
        name = (row_data.get('name') or '').strip()
        if sku:
            item = InventoryItem.objects.filter(
                organization=self.organization, sku__iexact=sku,
            ).first()
            if item:
                return item
        if name:
            candidates = list(InventoryItem.objects.filter(
                organization=self.organization, is_active=True,
            ).values('id', 'name'))
            names = [c['name'] for c in candidates]
            matches = difflib.get_close_matches(name, names, n=1, cutoff=0.75)
            if matches:
                target = next(c for c in candidates if c['name'] == matches[0])
                return InventoryItem.objects.get(pk=target['id'])
        return None

    # ──────────────────────────────────────────────────────────────
    # Main entry — parse (no DB writes)
    # ──────────────────────────────────────────────────────────────
    def parse(self, max_preview_rows=None, column_overrides=None) -> dict:
        rows = self._read_rows()
        if not rows:
            return {
                'rows': [], 'errors': [], 'column_map': {},
                'total_rows': 0, 'valid_rows': 0, 'error_rows': 0,
                'warnings': ['Empty file.'],
            }
        headers = rows[0]
        column_map = self.detect_columns(headers)
        if column_overrides:
            column_map.update(column_overrides)

        warnings = []
        required = {'sales': ['quantity'],
                    'purchase': ['quantity', 'unit_cost'],
                    'supplier_catalog': ['name']}
        for col in required.get(self.import_type, []):
            if col not in column_map:
                warnings.append(f'Required column "{col}" not detected in headers.')

        body = rows[1:]
        if max_preview_rows is not None:
            body = body[:max_preview_rows]

        parsed = []
        errors = []
        for offset, raw in enumerate(body):
            row_num = offset + 2  # 1-indexed + header row
            row_data = {}
            for std_col, idx in column_map.items():
                if idx < len(raw):
                    row_data[std_col] = raw[idx]

            row_errors = self.validate_row(row_data, row_num)
            # Normalize values
            normalized = {
                'row_num': row_num,
                'name': (row_data.get('name') or '').strip() if row_data.get('name') else '',
                'sku': (row_data.get('sku') or '').strip() if row_data.get('sku') else '',
                'quantity': _to_decimal(row_data.get('quantity')),
                'unit_cost': _to_decimal(row_data.get('unit_cost')) if row_data.get('unit_cost') else None,
                'movement_date': _parse_date(row_data.get('movement_date')),
                'supplier_name': (row_data.get('supplier_name') or '').strip(),
                'notes': (row_data.get('notes') or '').strip(),
                'errors': row_errors,
                'raw': raw,
            }
            if row_errors:
                errors.extend(row_errors)
            parsed.append(normalized)

        valid_rows = sum(1 for p in parsed if not p['errors'])
        return {
            'rows': parsed,
            'errors': errors,
            'column_map': column_map,
            'total_rows': len(parsed),
            'valid_rows': valid_rows,
            'error_rows': len(parsed) - valid_rows,
            'warnings': warnings,
            'headers': list(headers),
        }

    def preview(self, max_rows=10):
        return self.parse(max_preview_rows=max_rows)
